"""Report generation utilities."""

from __future__ import annotations

from typing import Any, Dict
from pathlib import Path
import os

from fpdf import FPDF
from openpyxl import Workbook
from uuid import uuid4

from GuideManager import GuideManager


class ReportGenerator:
    """Generates reports for quality-report methods from analyzed data."""

    def __init__(self, guide_manager: GuideManager) -> None:
        """Initialize with a ``GuideManager`` instance."""
        self.guide_manager = guide_manager

    def generate_template(self, method: str) -> Dict[str, Any]:
        """Return a report template for the given method."""
        return self.guide_manager.get_format(method)

    def generate(
        self,
        analysis: Dict[str, Any],
        complaint_info: Dict[str, str],
        output_dir: str | Path = ".",
    ) -> Dict[str, str]:
        """Create PDF and Excel reports from the analysis results.

        Parameters
        ----------
        analysis: Dict[str, Any]
            The analysis data returned from ``LLMAnalyzer``.
        complaint_info: Dict[str, str]
            Information about the complaint such as customer, subject and part code.
        output_dir: str | Path, optional
            Directory in which to save the generated files.

        Returns
        -------
        Dict[str, str]
            A dictionary containing the generated PDF and Excel file paths.
        """

        out_dir = Path(output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)

        unique_id = uuid4().hex
        pdf_path = out_dir / f"report_{unique_id}.pdf"
        excel_path = out_dir / f"report_{unique_id}.xlsx"

        # Create PDF
        pdf = FPDF()
        pdf.add_page()
        # Register a Unicode font for non-Latin characters
        env_font = os.getenv("FONT_PATH")
        if env_font:
            font_path = Path(env_font)
        else:
            font_path = Path(__file__).resolve().parents[1] / "Fonts" / "DejaVuSans.ttf"
        if not font_path.exists():
            fallback = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
            if fallback.exists():
                font_path = fallback
            else:
                raise FileNotFoundError(
                    "Font file not found. Checked "
                    f"{font_path} and {fallback}. Set FONT_PATH to override."
                )
        pdf.add_font("DejaVu", "", str(font_path), uni=True)
        pdf.set_font("DejaVu", size=12)
        pdf.cell(0, 10, txt="Analysis Report", ln=1)
        customer = complaint_info.get("customer", "")
        subject = complaint_info.get("subject", "")
        part_code = complaint_info.get("part_code", "")
        pdf.cell(0, 10, txt=f"Customer: {customer}", ln=1)
        pdf.cell(0, 10, txt=f"Subject: {subject}", ln=1)
        pdf.cell(0, 10, txt=f"Part Code: {part_code}", ln=1)
        pdf.ln(5)
        entries = []
        seen = set()
        for key, value in analysis.items():
            if key == "full_text" and "full_report" in analysis:
                continue
            response = value.get("response", "") if isinstance(value, dict) else str(value)
            if response in seen:
                continue
            seen.add(response)
            entries.append((key, response))

        for key, response in entries:
            line = f"{key}: {response}"
            width = getattr(pdf, "epw", 0)
            pdf.multi_cell(width, 10, txt=line)
        pdf.output(str(pdf_path))

        # Create Excel
        wb = Workbook()
        ws = wb.active
        ws.append(["Customer", customer])
        ws.append(["Subject", subject])
        ws.append(["Part Code", part_code])
        ws.append([])
        ws.append(["Step", "Response"])
        for key, response in entries:
            ws.append([key, response])
        wb.save(str(excel_path))

        return {"pdf": str(pdf_path), "excel": str(excel_path)}


__all__ = ["ReportGenerator"]
