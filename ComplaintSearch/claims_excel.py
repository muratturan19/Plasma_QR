"""Excel-based query utilities for customer claims."""

from __future__ import annotations

from typing import Any, Dict, Iterable, List, Tuple
from pathlib import Path
from datetime import datetime
import os
import logging

from dotenv import load_dotenv
from importlib import resources

from difflib import SequenceMatcher
from openpyxl import load_workbook

from . import normalize_text


class ExcelClaimsSearcher:
    """Search complaint records stored in an Excel file."""

    def __init__(self, path: str | Path | None = None) -> None:
        """Initialize with optional Excel file ``path``.

        When ``path`` is ``None``, ``CLAIMS_FILE_PATH`` is read from the
        ``.env`` file. If the variable is unset, the bundled ``claims.xlsx``
        inside the ``CC`` package is used.
        """
        if path is None:
            load_dotenv()
            path = os.getenv("CLAIMS_FILE_PATH")
            if path is None:
                path = resources.files("CC").joinpath("claims.xlsx")
        self.path = Path(path)

    def _load_headers(
        self, rows: Iterable[tuple[Any, ...]]
    ) -> Tuple[List[str], Dict[str, int]]:
        """Return header names and mapping from normalized keys to column index."""
        for row in rows:
            non_empty = [c for c in row if c not in (None, "")]
            if len(non_empty) >= 3:
                headers = [str(c) if c is not None else "" for c in row]
                mapping = {
                    normalize_text(h): idx for idx, h in enumerate(headers)
                }
                return headers, mapping
        return [], {}

    def search(
        self,
        filters: Dict[str, str],
        year: int | None = None,
        start_year: int | None = None,
        end_year: int | None = None,
    ) -> List[Dict[str, Any]]:
        """Return rows matching ``filters`` and optional year constraints.

        Parameters
        ----------
        filters:
            Mapping of field names to filter values. Keys are compared
            case-insensitively. The ``complaint`` filter performs a substring
            search while other fields must match exactly.
        year:
            Optional single-year constraint applied to a ``date`` column when
            present. Takes precedence over ``start_year``/``end_year``.
        start_year, end_year:
            Optional inclusive date range boundaries applied to a ``date``
            column.

        Returns
        -------
        List[Dict[str, Any]]
            Matching rows as dictionaries keyed by normalized headers.
        """
        if not self.path.exists():
            logging.warning("Excel file not found at %s", self.path)
            return []

        wb = load_workbook(self.path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers, indices = self._load_headers(rows)
        if not headers:
            wb.close()
            return []
        results: List[Dict[str, Any]] = []

        for row in rows:
            record = {
                key: row[idx] if idx < len(row) else None
                for key, idx in indices.items()
            }
            if "date" in indices:
                value = record.get("date")
                if isinstance(value, str):
                    try:
                        value = datetime.fromisoformat(value)
                    except ValueError:
                        continue
            else:
                value = None

            if year is not None and value is not None:
                if getattr(value, "year", None) != year:
                    continue
            elif value is not None:
                yr = getattr(value, "year", None)
                if start_year is not None and yr is not None and yr < start_year:
                    continue
                if end_year is not None and yr is not None and yr > end_year:
                    continue
            match = True
            for key, val in filters.items():
                if not val:
                    continue
                key = normalize_text(key)
                cell_raw = str(record.get(key, ""))
                cell = normalize_text(cell_raw)
                val_norm = normalize_text(str(val))
                if key == "complaint":
                    if val_norm in cell:
                        continue
                    ratio = SequenceMatcher(None, val_norm, cell).ratio()
                    if ratio < 0.8:
                        match = False
                        break
                else:
                    ratio = SequenceMatcher(None, val_norm, cell).ratio()
                    if cell != val_norm and ratio < 0.8:
                        match = False
                        break
            if match:
                results.append(record)

        wb.close()
        return results

    def unique_values(self, field: str) -> List[str]:
        """Return sorted unique values for ``field``.

        Parameters
        ----------
        field:
            Column name to extract unique values from. The name is
            normalized using :func:`normalize_text` before matching headers.

        Returns
        -------
        List[str]
            Unique cell values as strings. Empty cells are ignored.
        """
        if not self.path.exists():
            return []

        wb = load_workbook(self.path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers, indices = self._load_headers(rows)
        if not headers:
            wb.close()
            return []

        key = normalize_text(field)
        if key not in indices:
            wb.close()
            return []

        index = indices[key]
        values = set()
        for row in rows:
            if index >= len(row):
                continue
            val = row[index]
            if val is None:
                continue
            text = str(val).strip()
            if text:
                values.add(text)

        wb.close()
        return sorted(values)


__all__ = ["ExcelClaimsSearcher"]
