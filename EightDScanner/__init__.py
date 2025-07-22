"""Scan 8D report Excel files and store results in SQLite."""

from __future__ import annotations

import sqlite3
from pathlib import Path
from typing import List

from openpyxl import load_workbook


class EightDScanner:
    """Scan Excel reports for key fields and persist them."""

    def __init__(self, reports_dir: str | Path, db_path: str | Path = "eight_d.db") -> None:
        self.reports_dir = Path(reports_dir)
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        self.db_path = Path(db_path)
        self._init_db()

    def _init_db(self) -> None:
        with sqlite3.connect(self.db_path) as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS reports (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    material_code TEXT,
                    description TEXT,
                    customer TEXT,
                    root_cause TEXT,
                    permanent_action TEXT
                )
                """
            )

    @staticmethod
    def _normalize(text: str) -> str:
        return text.strip().lower().replace(" ", "")

    def _extract_rows(self, path: Path) -> List[tuple[str, str, str, str, str]]:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(c).strip() if c is not None else "" for c in next(rows, [])]
        mapping = {self._normalize(h): idx for idx, h in enumerate(headers)}
        keys = {
            "material_code": ["malzemekodu", "parcakodu"],
            "description": ["tanım", "tanim"],
            "customer": ["müşteri", "musteri"],
            "root_cause": ["kökneden", "kokneden"],
            "permanent_action": ["kalıcıaksiyon", "kaliciaksiyon"],
        }
        indices = {}
        for field, aliases in keys.items():
            for alias in aliases:
                if alias in mapping:
                    indices[field] = mapping[alias]
                    break
        results: List[tuple[str, str, str, str, str]] = []
        if len(indices) < 5:
            wb.close()
            return results
        for row in rows:
            try:
                values = [row[indices[f]] if indices[f] < len(row) else "" for f in indices]
                results.append(tuple(str(v).strip() if v is not None else "" for v in values))
            except Exception:
                continue
        wb.close()
        return results

    def scan(self) -> int:
        """Scan all Excel files and persist extracted rows."""
        count = 0
        for path in self.reports_dir.glob("*.xlsx"):
            rows = self._extract_rows(path)
            with sqlite3.connect(self.db_path) as conn:
                conn.executemany(
                    "INSERT INTO reports(material_code, description, customer, root_cause, permanent_action) VALUES (?, ?, ?, ?, ?)",
                    rows,
                )
                count += len(rows)
        return count

__all__ = ["EightDScanner"]
