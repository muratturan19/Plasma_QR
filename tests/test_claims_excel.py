import os
import tempfile
import unittest
from datetime import datetime
from unittest.mock import patch
from pathlib import Path

from ComplaintSearch import claims_excel
from ComplaintSearch.claims_excel import ExcelClaimsSearcher
from openpyxl import Workbook, load_workbook


class ExcelClaimsSearchTest(unittest.TestCase):
    """Tests for ExcelClaimsSearcher.search."""

    def _create_file(self, path: str, headers=None) -> None:
        wb = Workbook()
        ws = wb.active
        if headers is None:
            headers = ["complaint", "customer", "subject", "part_code", "date"]
        ws.append(headers)
        ws.append(["noise", "ACME", "engine", "X1", datetime(2023, 1, 1)])
        ws.append(["crack", "BETA", "body", "X2", datetime(2022, 5, 1)])
        wb.save(path)

    def test_search_filters(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "claims.xlsx")
            self._create_file(file_path)
            searcher = ExcelClaimsSearcher(file_path)
            result = searcher.search({"customer": "ACME"}, year=2023)
            self.assertEqual(len(result), 1)
            self.assertEqual(result[0]["customer"], "ACME")
            empty = searcher.search({"customer": "ACME"}, year=2022)
            self.assertEqual(empty, [])

    def test_normalization_and_fuzzy(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "claims.xlsx")
            self._create_file(file_path)

            wb = load_workbook(file_path)
            ws = wb.active
            ws.append(["\u015fikayet var", "GAMMA", "door", "X3", datetime(2023, 2, 1)])
            wb.save(file_path)

            searcher = ExcelClaimsSearcher(file_path)

            accent = searcher.search({"complaint": "sikayet"})
            self.assertEqual(len(accent), 1)
            self.assertEqual(accent[0]["customer"], "GAMMA")

            typo = searcher.search({"complaint": "noize"})
            self.assertEqual(len(typo), 1)
            self.assertEqual(typo[0]["customer"], "ACME")

    def test_turkish_headers(self) -> None:
        """Ensure search works when headers are Turkish."""
        headers = [
            "müşteri şikayeti",
            "müşteri",
            "konu",
            "parça kodu",
            "tarih",
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "claims.xlsx")
            self._create_file(file_path, headers=headers)
            searcher = ExcelClaimsSearcher(file_path)
            result = searcher.search({"customer": "ACME"}, year=2023)
            self.assertEqual(len(result), 1)
            self.assertIn("customer", result[0])
            self.assertEqual(result[0]["customer"], "ACME")

    def test_unique_values(self) -> None:
        """``unique_values`` should return sorted distinct entries."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "claims.xlsx")
            self._create_file(file_path)

            wb = load_workbook(file_path)
            ws = wb.active
            ws.append(["extra", "ACME", "engine", "X1", datetime(2024, 1, 1)])
            wb.save(file_path)

            searcher = ExcelClaimsSearcher(file_path)
            customers = searcher.unique_values("customer")
            self.assertEqual(customers, ["ACME", "BETA"])

    def test_env_default_path(self) -> None:
        """File path should come from ``CLAIMS_FILE_PATH`` when not provided."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "claims.xlsx")
            self._create_file(file_path)
            with patch.dict("os.environ", {"CLAIMS_FILE_PATH": file_path}):
                searcher = ExcelClaimsSearcher()
                result = searcher.search({"customer": "ACME"}, year=2023)
                self.assertEqual(len(result), 1)

    def test_load_dotenv_called(self) -> None:
        """``load_dotenv`` should be invoked when ``path`` is ``None``."""
        with patch.dict("os.environ", {"CLAIMS_FILE_PATH": "f"}):
            with patch.object(claims_excel, "load_dotenv") as mock_load:
                searcher = ExcelClaimsSearcher()
                self.assertTrue(mock_load.called)
                self.assertEqual(searcher.path, Path("f"))

    def test_bundled_file_outside_repo(self) -> None:
        """Bundled Excel file should load when run outside the repo root."""
        repo_root = Path(__file__).resolve().parents[1]
        expected = repo_root / "CC" / "claims.xlsx"
        with tempfile.TemporaryDirectory() as tmpdir:
            cwd = os.getcwd()
            try:
                os.chdir(tmpdir)
                with patch.dict("os.environ", {}, clear=True), \
                        patch.object(claims_excel, "load_dotenv"):
                    searcher = ExcelClaimsSearcher()
            finally:
                os.chdir(cwd)
        self.assertEqual(searcher.path, expected)
        self.assertTrue(len(searcher.search({})) > 0)

    def test_year_range(self) -> None:
        """Records should be filterable by a start and end year."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "claims.xlsx")
            self._create_file(file_path)
            searcher = ExcelClaimsSearcher(file_path)
            both = searcher.search({}, start_year=2022, end_year=2023)
            self.assertEqual(len(both), 2)
            just_2022 = searcher.search({}, start_year=2022, end_year=2022)
            self.assertEqual(len(just_2022), 1)
            overlap = searcher.search({}, year=2023, start_year=2022, end_year=2023)
            self.assertEqual(len(overlap), 1)


if __name__ == "__main__":
    unittest.main()
